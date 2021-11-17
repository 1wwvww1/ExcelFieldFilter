# -*- coding=utf-8 -*-

import time, logging as log
from common import LogUtil as logu, ProjectCommon as common
from pyexcel import ExcelExport as excel
from entity import ConfigEntity as config
import openpyxl

class ExcelService:

    excelPath = None
    excelInstance = None
    loadSheetName = None
    sheetInstance = None

    @logu.logExceptionHandler(log_func=logu.exceptionHandler)
    def __init__(self, excelPath):
        if excelPath is None and common.FileExist(excelPath):
            raise Exception("输入的Excel文件不存在 ( " + str(excelPath) + " )")
        else:
            self.excelPath = excelPath
        self.excelInstance = openpyxl.load_workbook(excelPath, data_only=True)

    def NotRangeException(self, message):
        return Exception("数据不符合表格范围数大小！ " + message)

    # @logu.logExceptionHandler(log_func=logu.exceptionHandler)
    def loadSheetName(self, sheetName):
        sheetNameSet = self.excelInstance.get_sheet_names()
        for sheet in sheetNameSet:

            if sheetName.__eq__(sheet):
                self.loadSheetName = sheet
                break

        if self.loadSheetName is not None and self.excelInstance is not None:
            self.sheetInstance = self.excelInstance.get_sheet_by_name(self.loadSheetName)
            return self.sheetInstance
        else:
            raise Exception("配置的sheet名称不存在于Excel表格中或excel操作的实例未定义")

    @logu.logExceptionHandler(log_func=logu.exceptionHandler, isContinue=False)
    def getSheetMaxRow(self, Sheet):
        i = Sheet.max_row
        real_max_row = 0
        while i > 0:
            row_dict = {i.value for i in Sheet[i]}
            if row_dict == {None}:
                i = i - 1
            else:
                real_max_row = i
                break
        return real_max_row

    @logu.logExceptionHandler(log_func=logu.exceptionHandler, isContinue=False)
    def getSheetMaxCell(self, Sheet):
        i = Sheet.max_column
        # real_max_cell = 0
        # sheet_max = i = Sheet.max_column
        # while i > 0:
        #     row_dict = {i.value for i in Sheet[i]}
        #     if row_dict == {None}:
        #         i = i - 1
        #     else:
        #         real_max_cell = i
        #         break
        # if sheet_max > real_max_cell:
        #     return sheet_max
        # 更换算法，如果最后列不为空就直接返回，否则最后列往前递增，只要拿到一列值存在就直接认为这是最大列
        while i>0:
            row_dict = Sheet.cell(row=1, column=i).value
            if row_dict == {None} or "None".__eq__(row_dict) :
                i = i-1
            else:
                return i

    # @logu.logExceptionHandler(log_func=logu.exceptionHandler)
    # def readExcelToMap(self, sheet, start_row, start_cell, max_row):
    #     resultMap = {}
    #     next_cell = chr(bytes(start_cell, 'utf-8')[0] + 1)
    #     for i in range(start_row, max_row + 1):
    #         try:
    #             key = (int(sheet[start_cell + str(i)].value))
    #             resultMap[key] = (sheet[next_cell + str(i)].value)
    #         except Exception as ex:
    #             raise Exception("获取excel列失败！具体错误为：" + str(ex))
    #     return resultMap
    #
    # @logu.logExceptionHandler(log_func=logu.exceptionHandler)
    # def writeDataToExcel(self, sheet, start_row, start_cell, writeValue):
    #     sheet.cell(row=start_row,column=ord(start_cell)-64,value=writeValue)
    #     self.excelInstance.save(self.excelPath)

    # 返回索引对应的列字典 key = 索引名称 value = 对应列
    @logu.logExceptionHandler(log_func=logu.exceptionHandler, isContinue=False)
    def readExcelIndexToMap(self, indexList, indexStartRow, sheet):
        resultMap = {}
        maxCell = self.getSheetMaxCell(sheet)
        for cell in range(1, maxCell+1):
            fieldValue = sheet.cell(row=indexStartRow, column=cell).value
            for indexName in indexList:
                if fieldValue is not None and indexName.__eq__(fieldValue):
                    resultMap[indexName] = cell
        return resultMap

    # key = 对应行， value = 内容
    @logu.logExceptionHandler(log_func=logu.exceptionHandler, isContinue=False)
    def readExcelCellToMap(self, cell, sheet):
        resultMap = {}
        maxRow = self.getSheetMaxRow(sheet)
        for row in range(1, maxRow+1):
            if row == config.excelIndexStartRow:
                continue
            resultMap[row] = sheet.cell(row=row, column=cell).value
        return resultMap

    # 根据索引读取对应一行数据
    @logu.logExceptionHandler(log_func=logu.exceptionHandler, isContinue=False)
    def readExcelOneRow(self, indexCellList, row, sheet):
        resultList = []
        for key in indexCellList:
            resultList.append(str(sheet.cell(row=row, column=indexCellList[key]).value).strip())
        return resultList

@logu.logExceptionHandler(log_func=logu.exceptionHandler, isContinue=False)
def createExcelBook(newFilePath, newSheetNames=[], indexList=[], newExcelDataList=[]):
    excelUtil = excel.ExcelUtil(newFilePath)
    new_sheet_list = newSheetNames
    excelUtil.add_sheet(new_sheet_list)

    for newSheetNameIndex in range(0, newSheetNames.__len__()):
        newSheetName = newSheetNames[newSheetNameIndex]

        # 先建立索引行
        for indexCell in range(1,len(indexList)+1):
            excelUtil.set_cell(
                newSheetName,
                col=indexCell,
                row=1,
                value=indexList[indexCell-1],
                cell_fill="yellow",
                cell_alignment="center",
                font_bold=True,
                font_size=13
            )
            excelUtil.set_col_weight(newSheetName, indexCell, 15)

        # 将数据写到excel表格中
        newExcelData = newExcelDataList[newSheetNameIndex]
        for excelData in range(0, newExcelData.__len__()):
            excelRowData = newExcelData[excelData]
            for cell in range(0, excelRowData.__len__()):
                cellV = excelRowData[cell]
                if "None".__eq__(cellV) or cellV is {None} or "YTSNVGH" in cellV:
                    cellV = ""
                excelUtil.set_cell(
                    newSheetName,
                    col=cell+1,
                    row=excelData+2,
                    value=cellV,
                    cell_alignment="center",
                    font_bold=False,
                    font_size=10
                )
        # 4. 保存
    excelUtil.save()